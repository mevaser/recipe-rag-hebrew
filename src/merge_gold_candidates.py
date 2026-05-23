from __future__ import annotations

import csv
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


PROJECT_ROOT = Path(__file__).resolve().parents[1]
EVAL_DIR = PROJECT_ROOT / "eval"
ARCHIVE_DIR = EVAL_DIR / "archive"
CANDIDATE_PATTERN = "gold_candidates_*.jsonl"
OUTPUT_JSONL_PATH = ARCHIVE_DIR / "gold_candidates_all.jsonl"
OUTPUT_CSV_PATH = ARCHIVE_DIR / "gold_candidates_review.csv"
OUTPUT_XLSX_PATH = ARCHIVE_DIR / "gold_candidates_review.xlsx"
REVIEW_COLUMNS = [
    "id",
    "review_status",
    "category",
    "question",
    "reference_answer",
    "must_cite_chunk_ids",
    "source",
]


def configure_stdout() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")


def find_candidate_files() -> list[Path]:
    return sorted(
        path
        for path in ARCHIVE_DIR.glob(CANDIDATE_PATTERN)
        if path.resolve() != OUTPUT_JSONL_PATH.resolve()
    )


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as input_file:
        for line_number, line in enumerate(input_file, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            try:
                row = json.loads(stripped)
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSON in {path} line {line_number}: {exc}") from exc
            if not isinstance(row, dict):
                raise ValueError(f"Expected JSON object in {path} line {line_number}.")
            rows.append(row)
    return rows


def duplicate_key(row: dict[str, Any]) -> tuple[str, tuple[str, ...]]:
    question = str(row.get("question", "")).strip()
    chunk_ids = row.get("must_cite_chunk_ids", [])
    if not isinstance(chunk_ids, list):
        chunk_ids = [str(chunk_ids)]
    return question, tuple(str(chunk_id) for chunk_id in chunk_ids)


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    normalized["review_status"] = normalized.get("review_status") or "needs_review"
    chunk_ids = normalized.get("must_cite_chunk_ids", [])
    if not isinstance(chunk_ids, list):
        chunk_ids = [str(chunk_ids)]
    normalized["must_cite_chunk_ids"] = [str(chunk_id) for chunk_id in chunk_ids]
    return normalized


def deduplicate_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    seen: set[tuple[str, tuple[str, ...]]] = set()
    merged: list[dict[str, Any]] = []
    duplicates_removed = 0

    for row in rows:
        normalized = normalize_row(row)
        key = duplicate_key(normalized)
        if key in seen:
            duplicates_removed += 1
            continue
        seen.add(key)
        merged.append(normalized)

    return merged, duplicates_removed


def write_jsonl(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as output_file:
        for row in rows:
            output_file.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_review_csv(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=REVIEW_COLUMNS)
        writer.writeheader()
        writer.writerows(build_review_rows(rows))


def build_review_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    review_rows: list[dict[str, Any]] = []
    for row_id, row in enumerate(rows, start=1):
        review_rows.append(
            {
                "id": row_id,
                "review_status": row.get("review_status", "needs_review"),
                "category": row.get("category", ""),
                "question": row.get("question", ""),
                "reference_answer": row.get("reference_answer", ""),
                "must_cite_chunk_ids": json.dumps(row.get("must_cite_chunk_ids", []), ensure_ascii=False),
                "source": row.get("source", ""),
            }
        )
    return review_rows


def write_review_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "gold_candidates"
    worksheet.append(REVIEW_COLUMNS)

    for review_row in build_review_rows(rows):
        worksheet.append([review_row[column] for column in REVIEW_COLUMNS])

    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    wrapped_columns = {"question", "reference_answer", "must_cite_chunk_ids", "source"}
    for column_index, column_name in enumerate(REVIEW_COLUMNS, start=1):
        max_length = len(column_name)
        for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for value_cell in cell:
                value = "" if value_cell.value is None else str(value_cell.value)
                max_length = max(max_length, min(len(value), 80))
                if column_name in wrapped_columns:
                    value_cell.alignment = Alignment(wrap_text=True, vertical="top")

        header_cell = worksheet.cell(row=1, column=column_index)
        header_cell.alignment = Alignment(vertical="top")
        width = min(max(max_length + 2, 10), 70)
        worksheet.column_dimensions[header_cell.column_letter].width = width

    workbook.save(path)


def print_summary(
    candidate_files_found: int,
    total_rows_loaded: int,
    duplicates_removed: int,
    final_rows_written: int,
    output_jsonl_path: Path,
    output_csv_path: Path,
    output_xlsx_path: Path,
) -> None:
    print("Gold candidate merge summary")
    print("============================")
    print(f"Candidate files found: {candidate_files_found}")
    print(f"Total rows loaded: {total_rows_loaded}")
    print(f"Duplicates removed: {duplicates_removed}")
    print(f"Final rows written: {final_rows_written}")
    print(f"Output JSONL path: {output_jsonl_path}")
    print(f"Output CSV path: {output_csv_path}")
    print(f"Output XLSX path: {output_xlsx_path}")


def main() -> None:
    configure_stdout()
    candidate_files = find_candidate_files()
    loaded_rows: list[dict[str, Any]] = []

    for path in candidate_files:
        loaded_rows.extend(load_jsonl(path))

    merged_rows, duplicates_removed = deduplicate_rows(loaded_rows)
    write_jsonl(merged_rows, OUTPUT_JSONL_PATH)
    write_review_csv(merged_rows, OUTPUT_CSV_PATH)
    write_review_xlsx(merged_rows, OUTPUT_XLSX_PATH)

    print_summary(
        candidate_files_found=len(candidate_files),
        total_rows_loaded=len(loaded_rows),
        duplicates_removed=duplicates_removed,
        final_rows_written=len(merged_rows),
        output_jsonl_path=OUTPUT_JSONL_PATH,
        output_csv_path=OUTPUT_CSV_PATH,
        output_xlsx_path=OUTPUT_XLSX_PATH,
    )


if __name__ == "__main__":
    main()
